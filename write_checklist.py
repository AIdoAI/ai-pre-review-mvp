#!/usr/bin/env python3
"""把预审结果写进《AI预审硬性材料核查表》xlsx（按现有 M1-M12 列与下拉选项）。

用法：
  python3.13 write_checklist.py --root "/路径/1-59" --xlsx "/路径/AI预审硬性材料核查表_1-59.xlsx" [--reviewer AI预审]

逐个批次读取 <root>/批次*/_预审结果/<参赛用户N>/rule_results.json，按编号匹配 xlsx 行，
只更新 M1-M12 状态/说明/严重程度 + 总体结论/需补/处理建议/审核人/状态，保留描述列(项目名/单位等)。
并打印每个用户"需人工核对"的清单。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)
    except Exception:
        pass

# M列 → 归属的规则 ID（前缀匹配）
M_RULES = {
    "M1": ["HR-3.1-APPLICATION"],
    "M2": ["HR-2.1-LICENSE"],
    "M3": ["HR-2.1-CREDIT"],
    "M4": ["HR-2.1-DECLARATION"],
    "M5": ["HR-2.1-COMMITMENT", "HR-6.1", "HR-6.3", "MR-COMMITMENT-CONTENT", "MR-COMMITMENT-VISUAL"],
    "M6": ["HR-2.1-FINANCIAL"],
    "M7": ["HR-2.1-RD-INVESTMENT"],
    "M8": ["HR-2.5-BUILDING", "HR-2.5-PLANNED"],
    "M9": ["HR-2.3-JOINT", "MR-JOINT-SUPPORT-MATERIAL", "HR-1.2-"],
    "M10": ["HR-2.4", "PARENT", "上级单位关联"],
    "M11": ["MR-PARSE-QUALITY", "MR-COMMITMENT-VISUAL"],
    "M12": ["MR-COMPANY-CONSISTENCY", "MR-ENTITY-TYPE"],
}
M_TITLES = {
    "M1": "任务书/申报书", "M2": "营业执照/主体资格", "M3": "信用记录证明",
    "M4": "法定代表人无重大违法声明", "M5": "真实性承诺书/签章", "M6": "主营收入/财务证明",
    "M7": "研发投入证明", "M8": "项目进展/实施条件证明", "M9": "联合体材料",
    "M10": "上级单位/授权材料", "M11": "签字盖章/扫描可读性", "M12": "单位/项目信息一致性",
}
# 状态选项：✓通过 / ✗缺失 / △需复核 / N/A不适用   严重程度：无/低/中/高
STATUS_PASS, STATUS_MISS, STATUS_REVIEW, STATUS_NA = "✓通过", "✗缺失", "△需复核", "N/A不适用"


def m_cell(rule_ids_present: list[dict], m_key: str) -> tuple[str, str, str]:
    """返回 (状态, 说明/需补, 严重程度)。"""
    prefixes = M_RULES[m_key]
    items = [r for r in rule_ids_present if any(p in r["rule_id"] for p in prefixes)]
    if not items:
        return STATUS_NA, "", "无"
    statuses = {r["status"] for r in items}
    nonpass = [r for r in items if r["status"] != "pass"]
    note = "；".join(re.sub(r"\s+", " ", r["reason"]).strip()[:60] for r in nonpass)[:240]
    if "fail" in statuses:
        return STATUS_MISS, note, "高"
    if "manual_review" in statuses:
        return STATUS_REVIEW, note, "中"
    if "not_assessable" in statuses:
        return STATUS_REVIEW, note or "解析不全/未覆盖，需人工核对", "低"
    return STATUS_PASS, "", "无"


OVERALL_TO_CONCL = {
    "预审不通过": "不通过", "待人工复核": "需人工复核",
    "建议补正": "需补充材料", "预审通过": "形式预审通过", "局部样本验证完成": "需人工复核",
}
CONCL_TO_ACTION = {
    "不通过": "退回", "需人工复核": "提交人工复核",
    "需补充材料": "补正后通过", "形式预审通过": "通过",
}


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--root", type=Path, required=True, help="批次父目录(含 批次*/_预审结果)")
    ap.add_argument("--xlsx", type=Path, required=True, help="核查表 xlsx")
    ap.add_argument("--reviewer", default="AI预审", help="审核人列填什么")
    ap.add_argument("--password", help="xlsx 打开密码(如加密)")
    args = ap.parse_args()

    import openpyxl
    if args.password:
        import io, msoffcrypto
        buf = io.BytesIO()
        with open(args.xlsx, "rb") as fh:
            off = msoffcrypto.OfficeFile(fh); off.load_key(password=args.password); off.decrypt(buf)
        wb = openpyxl.load_workbook(buf)
    else:
        wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["硬性材料核查表"]
    header = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]

    def col(name_exact: str) -> int:
        return header.index(name_exact) + 1

    c_no = col("参赛用户编号")
    row_by_no = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, c_no).value
        if v is not None and str(v).strip():
            row_by_no[str(v).strip()] = r

    # 收集各用户 rule_results
    written, todo_report = [], []
    for rr_path in sorted(args.root.glob("批次*/_预审结果/*/rule_results.json")):
        user = rr_path.parent.name  # 参赛用户N
        m = re.search(r"(\d+)", user)
        if not m:
            continue
        no = m.group(1)
        if no not in row_by_no:
            print(f"  ⚠️ {user}: 编号 {no} 不在核查表，跳过")
            continue
        results = json.loads(rr_path.read_text(encoding="utf-8")).get("results", [])
        overall = json.loads(rr_path.read_text(encoding="utf-8")).get("overall_status", "")
        row = row_by_no[no]
        manual_items = []
        for i, m_key in enumerate(M_RULES):  # M1..M12，每个占 3 列，从“M1 ...-状态”起
            base = col(f"{m_key} {M_TITLES[m_key]}-状态")
            status, note, sev = m_cell(results, m_key)
            ws.cell(row, base).value = status
            ws.cell(row, base + 1).value = note
            ws.cell(row, base + 2).value = sev
            if status in (STATUS_MISS, STATUS_REVIEW):
                manual_items.append(f"{m_key} {M_TITLES[m_key]}({status})")
        concl = OVERALL_TO_CONCL.get(overall, "需人工复核")
        need = "；".join(manual_items)
        ws.cell(row, col("总体结论")).value = concl
        ws.cell(row, col("需补材料汇总")).value = need
        ws.cell(row, col("处理建议")).value = CONCL_TO_ACTION.get(concl, "提交人工复核")
        ws.cell(row, col("审核人")).value = args.reviewer
        ws.cell(row, col("员工确认状态")).value = "待确认"
        written.append(no)
        if manual_items:
            todo_report.append((no, concl, manual_items))

    out = args.xlsx if not args.password else args.xlsx.with_name(args.xlsx.stem + "_已写入.xlsx")
    wb.save(out)
    print(f"\n已写入 {len(written)} 个用户 → {out}")
    print(f"编号：{', '.join(sorted(written, key=int))}")
    print("\n===== 需要你亲自核对的用户（AI 判为缺失/需复核的项）=====")
    for no, concl, items in sorted(todo_report, key=lambda x: int(x[0])):
        print(f"  用户{no} [{concl}]：{'；'.join(items)}")


if __name__ == "__main__":
    main()
