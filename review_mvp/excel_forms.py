"""从导出的参赛明细 Excel 自动生成各参赛单位的 form_answers。

对接 boss 指定的"Excel + 参赛用户文件夹"工作流：一行 = 一家，自动映射申报方式、
联合材料三选一、项目阶段、申报主体/联合成员，并按「提交项目任务书」文件名前缀的数字
匹配到 参赛用户{N} 文件夹。

依赖：openpyxl（读 .xlsx）；加密文件再需 msoffcrypto-tool。两者按需惰性导入，缺失给安装提示。
铁律：单位性质（国企/民企）不从 Excel 采信，留空 → 由营业执照/股权材料人工核实。
"""

from __future__ import annotations

import io
import re
from pathlib import Path
from typing import Any


def _pick_applicant_sheet(wb: Any):
    """选含参赛信息的 sheet（表头同时含"申报方式"和"单位名称"）；否则按名/退回首个。"""
    for ws in wb.worksheets:
        try:
            first = next(ws.iter_rows(values_only=True))
        except StopIteration:
            continue
        hdr = [("" if c is None else str(c).strip()) for c in first]
        if any("申报方式" in h for h in hdr) and any("单位名称" in h for h in hdr):
            return ws
    for name in ("申报主表", "主表"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.worksheets[0]


def _load_rows(path: Path, password: str | None) -> list[dict[str, str]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("读取 Excel 需要 openpyxl：pip install openpyxl") from exc
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        # 可能是加密 xlsx（CDFV2）→ 用 msoffcrypto 解密后再读
        try:
            import msoffcrypto
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("该 Excel 可能已加密，需要 msoffcrypto-tool：pip install msoffcrypto-tool") from exc
        if not password:
            raise RuntimeError("该 Excel 已加密，请用 --excel-password 提供打开密码")
        buf = io.BytesIO()
        with open(path, "rb") as fh:
            office = msoffcrypto.OfficeFile(fh)
            office.load_key(password=password)
            office.decrypt(buf)
        import openpyxl as _o
        wb = _o.load_workbook(buf, data_only=True, read_only=True)
    ws = _pick_applicant_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = ["" if c is None else str(c).strip() for c in rows[0]]
    out: list[dict[str, str]] = []
    for raw in rows[1:]:
        if not any(c is not None and str(c).strip() for c in raw):
            continue
        out.append({
            header[i]: ("" if i >= len(raw) or raw[i] is None else str(raw[i]).strip())
            for i in range(len(header))
        })
    return out


def _col(row: dict[str, str], name: str) -> str:
    """按完整表头名取值（容忍前后空格/全角差异，用包含匹配但用完整名避免歧义）。"""
    if name in row and row[name]:
        return row[name]
    for key, value in row.items():
        if name in key and value:
            return value
    return ""


def _partner_name(intro: str) -> str | None:
    """从联合单位简介文本里取开头的单位全称。"""
    m = re.match(r"\s*([一-龥（）()·]{4,40}(?:公司|集团|研究院|研究所|大学|学院|中心))", intro or "")
    return m.group(1) if m else None


def infer_material_type(cell: str) -> str | None:
    """从「关于联合体申报的相关材料」推断三选一类型；无/不明 → None（转人工）。"""
    text = (cell or "").strip()
    if not text or text in {"无", "无协议", "/", "-", "—"}:
        return None
    if "合作协议" in text:
        return "stamped_project_cooperation_agreement"
    if "联合申报协议" in text or "联合申请协议" in text:
        return "stamped_joint_declaration_agreement"
    if "声明" in text:
        return "stamped_lead_declaration"
    return None


def _folder_number(row: dict[str, str]) -> str | None:
    task_book = _col(row, "提交项目任务书")
    m = re.match(r"\D*(\d+)", task_book)
    return m.group(1) if m else None


def row_to_form(row: dict[str, str]) -> dict[str, Any]:
    """把一行映射成 form_answers（不含 mode）。"""
    way = _col(row, "申报方式")
    is_joint = "联合" in way
    indep_raw = _col(row, "是否独立法人")
    is_independent = ("是" in indep_raw) and ("否" not in indep_raw)
    stage_raw = _col(row, "项目当前进展")
    stage = "building" if "正在建设" in stage_raw else ("planned" if "计划实施" in stage_raw else "other")
    lead_name = _col(row, "单位名称") or None

    if not is_joint:
        return {
            "is_joint_declaration": False,
            "project_stage": stage,
            "applicants": [
                {"entity_id": "E01", "entity_name": lead_name, "is_independent_legal_person": is_independent}
            ],
        }

    members = []
    for col in ("联合申报单位简介（之一）", "联合申报单位简介（之二）"):
        name = _partner_name(_col(row, col))
        if name:
            members.append(name)
    applicants = [{"entity_id": "E01", "entity_name": lead_name, "is_lead": True,
                   "is_independent_legal_person": is_independent}]
    for i, name in enumerate(members, start=2):
        applicants.append({"entity_id": f"E{i:02d}", "entity_name": name, "is_lead": False,
                           "is_independent_legal_person": True})
    if len(applicants) < 2:
        # 联合申报至少 2 家；成员名可能在协议里而非 Excel → 占位，交叉检查仍会提示
        applicants.append({"entity_id": "E02", "entity_name": "联合成员（见协议/待核）",
                           "is_lead": False, "is_independent_legal_person": True})
    return {
        "is_joint_declaration": True,
        "joint_declaration_material_type": infer_material_type(_col(row, "关于联合体申报的相关材料")),
        "project_stage": stage,
        "applicants": applicants,
    }


def build_forms_from_excel(
    excel_path: Path, parent_dir: Path, password: str | None = None, mode: str = "complete",
) -> tuple[dict[str, tuple[dict[str, Any], str]], dict[str, dict[str, Any]]]:
    """读 Excel → 返回 ({文件夹名: (form_answers, mode)}, {编号: form_answers})。

    按「提交项目任务书」文件名前缀数字 ↔ 参赛用户{N} 文件夹名里的数字 匹配。
    """
    by_number: dict[str, dict[str, Any]] = {}
    for row in _load_rows(Path(excel_path), password):
        number = _folder_number(row)
        if number:
            by_number[number] = row_to_form(row)

    folder_forms: dict[str, tuple[dict[str, Any], str]] = {}
    for sub in sorted(p for p in Path(parent_dir).iterdir() if p.is_dir()):
        m = re.search(r"(\d+)", sub.name)
        if m and m.group(1) in by_number:
            folder_forms[sub.name] = (by_number[m.group(1)], mode)
    return folder_forms, by_number
